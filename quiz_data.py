"""
从 品牌整理.xlsx 读取品牌数据，生成 quiz_data.js
运行: python quiz_data.py
"""

import openpyxl
import json
import random
import re

# U+2019 右单引号（L'Oréal 在 Excel 中用的是弯引号）
_curly = '\u2019'

# ============================================================
# 1. 读取 Excel
# ============================================================

EXCEL = r'C:\Users\J nash\Desktop\workbuddy\品牌\品牌整理.xlsx'

# 分类名归一化映射（其他表短名 → 重点表标准名）
CAT_MAP = {
    '时装': '时装与皮具',
    '皮具': '时装与皮具',
    '化妆品': '化妆品与香水',
    '香水': '化妆品与香水',
    '腕表': '腕表与珠宝',
    '珠宝': '腕表与珠宝',
    '眼镜': '眼镜',
    '酒类': '酒类',
    '精品零售': '精品零售',
    '其他': '其他',
    '陶瓷': '其他',
    '时装与皮具': '时装与皮具',
    '化妆品与香水': '化妆品与香水',
    '腕表与珠宝': '腕表与珠宝',
}

# 标准分类码
TYPE_CODE = {
    '时装与皮具': 'f',
    '化妆品与香水': 'c',
    '腕表与珠宝': 'w',
    '眼镜': 'g',
    '酒类': 'l',
    '其他': 'o',
    '精品零售': 'o',
}

# 特殊分类修正（品牌名关键词 → 正确分类）
SPECIAL_CAT = {
    # 开云眼镜（Kering 开云有独立眼镜部门）
    'Barton Perreira': 'g',   # 开云眼镜
    'Vuarnet': 'g',            # 开云眼镜
    'Kering Eyecatcher': 'g',
}

# 集团名归一化（去除序号前缀如"六、"等）
def normalize_group(name):
    if not name:
        return None
    name = name.strip()
    # 去掉前缀序号
    name = re.sub(r'^[一二三四五六七八九十百千万\d．、\s]+', '', name)
    return name

def normalize_cat(cat):
    if not cat:
        return None
    cat = cat.strip()
    return CAT_MAP.get(cat, cat)

def read_excel():
    wb = openpyxl.load_workbook(EXCEL)
    brands = []  # [(品牌, 集团, 分类, 类型码)]

    for sh in wb.sheetnames:
        ws = wb[sh]
        cur_group = None
        cur_cat = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            g_raw = row[0]
            c_raw = row[1]
            b_raw = row[2]

            g_str = str(g_raw).strip() if g_raw is not None else None
            c_str = str(c_raw).strip() if c_raw is not None else None
            b_str = str(b_raw).strip() if b_raw is not None else None

            if g_str and g_str != '集团':
                cur_group = normalize_group(g_str)
            if c_str and c_str not in ('分类', 'None'):
                cur_cat = normalize_cat(c_str)
            if b_str and b_str not in ('品牌', 'None'):
                brand = b_str
                group = cur_group
                cat = cur_cat

                # 特殊分类修正
                type_code = SPECIAL_CAT.get(brand, TYPE_CODE.get(cat, 'f'))

                if brand and group and cat:
                    brands.append((brand, group, cat, type_code))

    return brands

# ============================================================
# 2. 去重与合并（同品牌同集团去重，不同集团保留）
# ============================================================

def dedup(brands):
    seen = {}
    result = []
    for b, g, cat, tc in brands:
        key = (b, g)
        if key not in seen:
            seen[key] = True
            result.append((b, g, cat, tc))
    return result

# ============================================================
# 3. 集团内部迁移（将品牌从 A 集团移到 B 集团）
# ============================================================
# 格式: (品牌名, 原集团/空, 新集团, 新分类, 新类型码)
GROUP_MOVES = [
    # Sephora 属于 LVMH（美妆零售，不是精品零售）
    ("Sephora丝芙兰", "LVMH路威酩轩集团", "LVMH路威酩轩集团", "其他", "o"),

    # Rimowa 属于 LVMH（时装与皮具）
    ("Rimowa日默瓦", "LVMH路威酩轩集团", "LVMH路威酩轩集团", "时装与皮具", "f"),

    # Hugo Boss 香水属于 L'Oréal
    ("Hugo Boss波士香水", "Hugo Boss雨果博斯集团", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),
]

# 手动新增条目（在任何集团下）
MANUAL_ADDITIONS = [
    # Sephora 属于 LVMH（GROUP_MOVES 会删除，需要补回）
    ("Sephora丝芙兰", "LVMH路威酩轩集团", "其他", "o"),

    # Viktor & Rolf 拆成两条（使用 Excel 中的实际名称）
    ("Viktor&Rolf维果罗夫美妆", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),
    ("Viktor & Rolf维果罗夫", "OTB (Only The Brave)集团", "时装与皮具", "f"),

    # Bottega Veneta 香水属于 L'Oréal
    ("Bottega Veneta葆蝶家香水", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Balenciaga 香水属于 L'Oréal
    ("Balenciaga巴黎世家香水", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Alexander McQueen 香水属于 L'Oréal
    ("Alexander McQueen麦昆香水", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Creed 属于 L'Oréal
    ("Creed恺芮得", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Pomellato 香水属于 L'Oréal
    ("Pomellato宝曼兰朵香水", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Qeelin 香水属于 L'Oréal
    ("Qeelin麒麟香水", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Miu Miu 香水 → L'Oréal 欧莱雅（不是 Prada）
    ("Miu Miu缪缪美妆", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Valentino Beauty → L'Oréal 欧莱雅
    ("Valentino Beauty华伦天奴美妆", f"L{_curly}Oréal欧莱雅集团", "化妆品与香水", "c"),

    # Kering Eyecatcher → 开云眼镜
    ("Kering Eyecatcher", "Kering开云集团", "眼镜", "g"),
]

# ============================================================
# 4. 集团名归一化（显示用）
# ============================================================

GROUP_DISPLAY = {
    "LVMH路威酩轩集团": "LVMH 路威酩轩",
    "Kering开云集团": "Kering 开云",
    "Richemont历峰集团": "Richemont 历峰",
    "ESTēE LAUDER雅诗兰黛集团": "EL 雅诗兰黛",
    f"L{_curly}Oréal欧莱雅集团": "LOreal 欧莱雅",
    "六．Swatch斯沃琪集团": "Swatch 斯沃琪",
    "Swatch斯沃琪集团": "Swatch 斯沃琪",
    "Hugo Boss雨果博斯集团": "Hugo Boss 雨果博斯",
    "OTB (Only The Brave)集团": "OTB 集团",
    "Valentino 华伦天奴集团": "Valentino 华伦天奴",
    "Prada Group普拉达集团": "Prada Group 普拉达",
    "六．L'Oréal欧莱雅集团": "LOreal 欧莱雅",  # 归一化后残留
    "OTB (Only The Brave)集团": "OTB 集团",
    "其他品牌": "其他品牌",
    "ANTA Sports安踏": "ANTA 安踏",
    "Moncler盟可睐集团": "Moncler 盟可睐",
    "Prada Group普拉达集团": "Prada Group 普拉达",
    "Puig普伊格集团": "Puig 普伊格",
    "Rolex劳力士集团": "Rolex 劳力士",
    "Tapestry泰佩思琦集团": "Tapestry 泰佩思琦",
    "Salvatore Ferragamo菲拉格慕集团": "Salvatore Ferragamo 菲拉格慕",
    "TOD'S托德斯集团": "TOD'S 托德斯",
    "Valentino 华伦天奴集团": "Valentino 华伦天奴",
}

def display_group(g):
    return GROUP_DISPLAY.get(g, g)

# ============================================================
# 5. 生成 quiz_data.js
# ============================================================

def generate():
    raw = read_excel()
    print(f"读取到 {len(raw)} 条原始记录")

    # 应用集团迁移
    for b_name, old_g, new_g, new_cat, new_tc in GROUP_MOVES:
        # 从 raw 中移除该品牌（如果指定了原集团，只移除匹配项；否则移除所有同名品牌）
        if old_g:
            raw = [(b, g, c, tc) for b, g, c, tc in raw
                   if not (b == b_name and g == old_g)]
        else:
            raw = [(b, g, c, tc) for b, g, c, tc in raw if b != b_name]

    print(f"迁移后 {len(raw)} 条")

    # 合并：raw + 新增条目
    all_brands = raw + MANUAL_ADDITIONS

    # 去重
    brands = dedup(all_brands)
    print(f"去重后 {len(brands)} 条")

    # 生成 JS
    # 格式: [品牌中文名, 品牌英文名, 集团简称(选项用), 集团内部名, 类型码]
    lines = [
        "// 由 quiz_data.py 自动生成",
        f"// 生成时间: 2026-04-25",
        "",
        "// 格式: [品牌中文名, 品牌英文名, 集团简称(选项用), 集团内部名, 类型码]",
        "// 类型码: f=时装与皮具, c=化妆品与香水, w=腕表与珠宝, g=眼镜, l=酒类, o=其他",
        "",
        "const BRANDS = [",
    ]

    for b, g, cat, tc in brands:
        bn = b.replace('\\', '\\\\').replace('"', '\\"')
        gn = display_group(g).replace('\\', '\\\\').replace('"', '\\"')
        # 从品牌中文名中提取英文名（如 "Louis Vuitton路易威登" → "Louis Vuitton"）
        en_name = re.sub(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef\u3400-\u4dbf]+.*', '', bn).strip()
        # 如果品牌名全中文或全英文，则英文名为空
        if not en_name or en_name == bn:
            en_name = ''
        lines.append(f'  ["{bn}", "{en_name}", "{gn}", "{g}", "{tc}"],')

    lines.append("];")
    lines.append("")
    lines.append("// 统计")
    tcnt = {}
    for b, g, cat, tc in brands:
        tcnt[tc] = tcnt.get(tc, 0) + 1
    for k, v in sorted(tcnt.items()):
        print(f"  {k}: {v} 个")

    js = "\n".join(lines)
    out = r'C:\Users\J nash\Desktop\workbuddy\品牌\quiz_data.js'
    with open(out, "w", encoding="utf-8") as f:
        f.write(js)
    print(f"\n已生成 {out}")

    # 也输出汇总供核对
    by_group = {}
    for b, g, cat, tc in brands:
        dg = display_group(g)
        by_group.setdefault(dg, []).append(b)

    print("\n--- 按集团汇总 ---")
    for grp in sorted(by_group.keys()):
        brands_list = sorted(by_group[grp])
        print(f"  [{grp}] ({len(brands_list)}): {', '.join(brands_list[:5])}{'...' if len(brands_list) > 5 else ''}")

if __name__ == "__main__":
    generate()
